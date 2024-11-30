from utils import execute_query
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment


# История транзакций
def transaction_history(start_date=None, end_date=None, type=None):
    query = """
        SELECT i.ingredient_name, st.transaction_date, st.transaction_type, st.quantity, st.reason
        FROM stock_transactions st
        JOIN ingredients i ON st.ingredient_id = i.ingredient_id
    """

    conditions = []
    if start_date:
        conditions.append("st.transaction_date >= %s")
    if end_date:
        conditions.append("st.transaction_date <= %s")
    if type:
        conditions.append("st.transaction_type = %s")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY st.transaction_date DESC"

    params = []
    if start_date:
        params.append(start_date)
    if end_date:
        params.append(end_date)
    if type:
        params.append(type)

    transactions = execute_query(query, tuple(params))

    # Создание Excel-файла
    file_path = "transaction_history.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История транзакций"

    # Заголовки
    headers = ["Ингредиент", "Дата", "Тип", "Количество", "Причина"]
    sheet.append(headers)

    # Выравнивание заголовков по центру
    for col in sheet[1]:
        col.alignment = Alignment(horizontal="center", vertical="center")

    # Добавление данных
    if transactions:
        for transaction in transactions:
            ingredient_name, date, trans_type, quantity, reason = transaction

            # Добавляем строку данных в таблицу
            sheet.append([ingredient_name, date, trans_type, quantity, reason])

    # Форматирование столбца с датами
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=sheet.max_row):  # Столбец с датой — 2-й
        cell = row[0]
        cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Автоматическая настройка ширины столбцов
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Получаем букву столбца
        for cell in column:
            try:
                # Получение длины значения в ячейке
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Устанавливаем ширину столбца
        adjusted_width = max_length + 2  # Добавляем немного места для эстетики
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Сохранение файла
    workbook.save(file_path)

    return transactions, file_path


def items_transaction_history(start_date=None, end_date=None, type=None):
    query = """
        SELECT hi.item_name, sth.transaction_date, sth.transaction_type, sth.quantity, sth.reason
        FROM stock_transaction_houseitems sth
        JOIN household_items hi ON sth.item_id = hi.item_id
    """

    conditions = []
    if start_date:
        conditions.append("sth.transaction_date >= %s")
    if end_date:
        conditions.append("sth.transaction_date <= %s")
    if type:
        conditions.append("sth.transaction_type = %s")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY sth.transaction_date DESC"

    params = []
    if start_date:
        params.append(start_date)
    if end_date:
        params.append(end_date)
    if type:
        params.append(type)

    transactions = execute_query(query, tuple(params))

    # Создание Excel-файла
    file_path = "items_transaction_history.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История транзакций"

    # Заголовки
    headers = ["Товар", "Дата", "Тип", "Количество", "Причина"]
    sheet.append(headers)

    # Выравнивание заголовков по центру
    for col in sheet[1]:
        col.alignment = Alignment(horizontal="center", vertical="center")

    # Добавление данных
    if transactions:
        for transaction in transactions:
            item_name, date, trans_type, quantity, reason = transaction

            # Добавляем строку данных в таблицу
            sheet.append([item_name, date, trans_type, quantity, reason])

    # Форматирование столбца с датами
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=sheet.max_row):  # Столбец с датой — 2-й
        cell = row[0]
        cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Автоматическая настройка ширины столбцов
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Получаем букву столбца
        for cell in column:
            try:
                # Получение длины значения в ячейке
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Устанавливаем ширину столбца
        adjusted_width = max_length + 2  # Добавляем немного места для эстетики
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Сохранение файла
    workbook.save(file_path)

    return transactions, file_path




# Удаление транзакций
def delete_transaction(transaction_id):
    # Проверка, существует ли транзакция
    transaction = execute_query(
        "SELECT transaction_id FROM stock_transactions WHERE transaction_id = %s",
        (transaction_id,),
        fetchone=True
    )

    if transaction is None:
        print(f"Транзакция с ID {transaction_id} не найдена.")
        return

    # Удаление транзакции
    execute_query(
        "DELETE FROM stock_transactions WHERE transaction_id = %s",
        (transaction_id,),
        commit=True
    )

    print(f"Транзакция с ID {transaction_id} успешно удалена.")

